from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote
import json
import logging

import pandas as pd

from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, StreamingResponse
from sqlalchemy import inspect, text, func
from sqlalchemy.orm import Session

from .config import (
    ALPHA_VANTAGE_SYMBOL_MAP,
    FRED_US_INDICES,
    FRED_US_INDEX_NAMES,
    get_default_indices,
    WECHAT_NOTIFY_DATA_UPDATE_DEFAULT,
)
from .database import get_db as base_get_db, init_db, engine, SessionLocal, TZ_SHANGHAI, now_shanghai
from .fetcher import (
    backfill_all_history,
    fetch_alphavantage_latest,
    fetch_fred_latest_observation,
    fetch_history_for_index,
    fetch_history_for_indices,
    recalculate_moving_averages_and_diffs,
)
from .backtest import (
    backtest_dca,
    backtest_smart_dca,
    backtest_mean_reversion,
    backtest_mean_reversion_multi,
    backtest_trend_following,
    backtest_ma_diff,
    compute_metrics,
)
from .wechat import get_subscriber_list, send_custom_message
from .emailer import test_smtp_connection, send_email
from .models import Index, IndexPrice, DbAuditLog, SyncJobLog, NotificationSetting, LiveStrategy, Contact, EmailConfig
from .schemas import (
    IndexCreate,
    IndexUpdate,
    IndexRead,
    PriceQueryResponse,
    IndexPriceRead,
    TableInfo,
    TableDataResponse,
    DbAuditLogRead,
    AuditLogListResponse,
    SyncJobLogRead,
    SyncJobLogListResponse,
    TableRowUpdateRequest,
    IndexDateStatus,
    UsIndexQuoteRead,
    CalendarDataResponse,
    CalendarDataPoint,
    NotificationSettingRead,
    NotificationSettingUpdate,
    WechatSubscriberRead,
    SendCustomMessageRequest,
    BacktestRequest,
    BacktestResponse,
    BacktestPointRead,
    BacktestTradeRead,
    BacktestMetrics,
    LiveStrategyCreate,
    LiveStrategyRead,
    LiveStrategyUpdate,
    LiveStrategyFromBacktest,
    ContactCreate,
    ContactRead,
    ContactUpdate,
    EmailConfigRead,
    EmailConfigUpdate,
    MonitorDailyChangeItem,
    MonitorRiskAlertItem,
    MonitorStrategyTodoItem,
    MonitorDailySummary,
    MonitorDashboard,
)
from .scheduler import start_scheduler, shutdown_scheduler, run_live_strategies, _evaluate_live_signal

app = FastAPI(title="基金指数分析系统")
_log = logging.getLogger("fund_analyzer.main")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()
    db = next(base_get_db())
    try:
        # 初始化默认指数（如不存在）
        existing_codes = {obj.code for obj in db.query(Index).all()}
        for code in get_default_indices():
            if code not in existing_codes:
                db.add(Index(code=code, name=code))
        db.commit()
    finally:
        db.close()

    # 启动每日定时任务
    start_scheduler()


@app.on_event("shutdown")
def on_shutdown():
    shutdown_scheduler()


def get_db(request: Request):
    """
    带操作人标记的 Session 依赖：
    - 普通 HTTP 请求：operator = "api"
    - 若前端通过 Header 传 X-Operator，则优先使用该值
    """
    db: Session = SessionLocal()
    db.info["operator"] = request.headers.get("X-Operator", "api")
    try:
        yield db
    finally:
        db.close()


@app.post("/indices", response_model=IndexRead)
def add_index(index_in: IndexCreate, db: Session = Depends(get_db)):
    if not index_in.name or not index_in.name.strip():
        raise HTTPException(status_code=400, detail="指数名称不能为空")
    
    exists_code = db.query(Index).filter(Index.code == index_in.code).one_or_none()
    if exists_code:
        raise HTTPException(status_code=400, detail="该指数代码已存在")
    
    exists_name = db.query(Index).filter(Index.name == index_in.name).one_or_none()
    if exists_name:
        raise HTTPException(status_code=400, detail="该指数名称已存在")
    
    obj = Index(
        code=index_in.code.strip(),
        name=index_in.name.strip(),
        start_date=index_in.start_date,
        source=index_in.source.strip() if index_in.source else None,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)

    # 新增指数后自动拉取历史数据（有 start_date 则从该日至今，否则从一年前至今）
    try:
        today = datetime.now(TZ_SHANGHAI).date()
        start = index_in.start_date if index_in.start_date else today - timedelta(days=365)
        fetch_history_for_index(db, obj, start=start, end=today)
    except Exception:
        pass  # 拉取失败不影响创建成功，后续可由定时任务或手动同步补全

    return obj


@app.patch("/indices/{index_id}", response_model=IndexRead)
def update_index(index_id: int, index_in: IndexUpdate, db: Session = Depends(get_db)):
    obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="指数不存在")

    if index_in.start_date is not None:
        obj.start_date = index_in.start_date
    if index_in.source is not None:
        obj.source = index_in.source.strip() if index_in.source else None

    db.commit()
    db.refresh(obj)
    return obj


@app.get("/indices", response_model=List[IndexRead])
def list_indices(db: Session = Depends(get_db)):
    return db.query(Index).order_by(Index.id.asc()).all()


@app.get("/monitor/dashboard", response_model=MonitorDashboard)
def monitor_dashboard(db: Session = Depends(get_db)):
    latest_date = db.query(func.max(IndexPrice.trade_date)).scalar()
    if not latest_date:
        summary = MonitorDailySummary(
            date=None,
            total_indices=0,
            up_count=0,
            down_count=0,
            flat_count=0,
            top_gainers=[],
            top_losers=[],
        )
        return MonitorDashboard(
            date=None,
            daily_summary=summary,
            risk_alerts=[],
            strategy_todos=[],
        )

    rows = (
        db.query(IndexPrice)
        .filter(IndexPrice.trade_date == latest_date)
        .order_by(IndexPrice.index_code.asc())
        .all()
    )

    up = down = flat = 0
    daily_items: list[MonitorDailyChangeItem] = []
    for row in rows:
        change = row.change_pct if row.change_pct is not None else 0.0
        if change > 0:
            up += 1
        elif change < 0:
            down += 1
        else:
            flat += 1
        daily_items.append(
            MonitorDailyChangeItem(
                index_code=row.index_code,
                index_name=row.index_name,
                close=row.close,
                change_pct=row.change_pct,
            )
        )

    gainers = sorted(
        [i for i in daily_items if i.change_pct is not None],
        key=lambda x: x.change_pct,
        reverse=True,
    )[:5]
    losers = sorted(
        [i for i in daily_items if i.change_pct is not None],
        key=lambda x: x.change_pct,
    )[:5]

    risk_alerts: list[MonitorRiskAlertItem] = []
    for row in rows:
        reasons = []
        if row.change_pct is not None and row.change_pct <= -3:
            reasons.append("日跌幅<=-3%")
        if row.ma60_diff is not None and row.ma60_diff <= -100:
            reasons.append("MA60偏差<=-100")
        if row.ma120_diff is not None and row.ma120_diff <= -150:
            reasons.append("MA120偏差<=-150")
        if reasons:
            risk_alerts.append(
                MonitorRiskAlertItem(
                    index_code=row.index_code,
                    index_name=row.index_name,
                    reason="; ".join(reasons),
                    change_pct=row.change_pct,
                    ma60_diff=row.ma60_diff,
                )
            )

    # 策略待处理事项：展示最新交易日触发信号的策略
    strategy_todos: list[MonitorStrategyTodoItem] = []
    strategies = db.query(LiveStrategy).filter(LiveStrategy.enabled.is_(True)).all()
    for strat in strategies:
        latest_row = (
            db.query(IndexPrice)
            .filter(IndexPrice.index_code == strat.index_code)
            .order_by(IndexPrice.trade_date.desc())
            .first()
        )
        if not latest_row:
            continue
        try:
            params = json.loads(strat.params) if strat.params else {}
        except Exception:
            params = {}
        action, detail, _debug = _evaluate_live_signal(db, strat, params, latest_row)
        if action in ("buy", "sell"):
            strategy_todos.append(
                MonitorStrategyTodoItem(
                    strategy_id=strat.id,
                    name=strat.name,
                    index_code=strat.index_code,
                    index_name=strat.index_name,
                    action=action,
                    detail=detail,
                )
            )

    summary = MonitorDailySummary(
        date=latest_date.isoformat() if latest_date else None,
        total_indices=len(rows),
        up_count=up,
        down_count=down,
        flat_count=flat,
        top_gainers=gainers,
        top_losers=losers,
    )

    return MonitorDashboard(
        date=latest_date.isoformat() if latest_date else None,
        daily_summary=summary,
        risk_alerts=risk_alerts,
        strategy_todos=strategy_todos,
    )


@app.get("/us-indices/quote", response_model=List[UsIndexQuoteRead])
def get_us_indices_quote():
    """
    获取美股指数最新净值（纳斯达克、标普500、道琼斯等）。
    优先使用 FRED（FRED_API_KEY），无数据时尝试 Alpha Vantage（ALPHA_VANTAGE_API_KEY）。
    """
    result: List[UsIndexQuoteRead] = []
    for series_id in FRED_US_INDICES:
        obs = fetch_fred_latest_observation(series_id)
        if obs is None and series_id in ALPHA_VANTAGE_SYMBOL_MAP:
            obs = fetch_alphavantage_latest(ALPHA_VANTAGE_SYMBOL_MAP[series_id])
        if obs:
            name = FRED_US_INDEX_NAMES.get(series_id, series_id)
            result.append(
                UsIndexQuoteRead(
                    series_id=series_id,
                    name=name,
                    trade_date=obs["trade_date"],
                    close=obs["close"],
                )
            )
    return result


@app.delete("/indices/{index_id}")
def delete_index(
    index_id: int,
    delete_related: bool = False,
    db: Session = Depends(get_db),
):
    """
    删除指数。delete_related=True 时，会先删除 index_prices 中该指数的所有行情数据，再删除指数。
    """
    obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    deleted_prices = 0
    if delete_related:
        deleted_prices = db.query(IndexPrice).filter(IndexPrice.index_code == obj.code).delete()
        db.commit()
    db.delete(obj)
    db.commit()
    return {"ok": True, "deleted_prices": deleted_prices}


@app.post("/indices/sync_all_history")
def sync_all_history(db: Session = Depends(get_db)):
    """
    首次运行时调用：为当前所有指数一次性拉取历史数据并计算均线。
    可以重复调用，重复日期会被更新而不是重复插入。
    """
    indices = db.query(Index).all()
    if not indices:
        return {"updated": {}}

    log = SyncJobLog(job_type="sync_all_history", status="running")
    db.add(log)
    db.commit()
    db.refresh(log)

    try:
        updated, sources = fetch_history_for_indices(db, indices, start=None, end=None)
        log.status = "success"
        log.total_indices = len(indices)
        log.total_rows = int(sum(updated.values()))
        # 统计各数据源使用次数
        source_counts: dict[str, int] = {}
        for src in sources.values():
            if not src:
                continue
            source_counts[src] = source_counts.get(src, 0) + 1
        if source_counts:
            # 例如：csindex: 3
            parts = [f"{k}: {v}" for k, v in source_counts.items()]
            log.source_summary = ", ".join(parts)
        
        # 详情格式：indices code：开始日期-截止日期
        detail_parts = []
        for idx in indices:
            if idx.id in updated and updated[idx.id] > 0:
                # 获取该指数的日期范围
                from .fetcher import _get_last_trade_date
                last_date = _get_last_trade_date(db, idx.code)
                first_date = None
                if last_date:
                    # 查询最早日期
                    from sqlalchemy import select, func
                    from .models import IndexPrice
                    stmt = (
                        select(func.min(IndexPrice.trade_date))
                        .where(IndexPrice.index_code == idx.code)
                    )
                    first_date = db.execute(stmt).scalar_one_or_none()
                
                if first_date and last_date:
                    detail_parts.append(f"{idx.code}：{first_date.isoformat()}-{last_date.isoformat()}")
                elif last_date:
                    detail_parts.append(f"{idx.code}：-{last_date.isoformat()}")
        
        log.detail = "；".join(detail_parts) if detail_parts else json.dumps(updated, ensure_ascii=False)
        log.finished_at = now_shanghai()
        db.commit()
    except Exception as exc:
        db.rollback()
        log.status = "error"
        log.detail = f"{type(exc).__name__}: {exc}"
        log.finished_at = now_shanghai()
        db.commit()
        raise

    return {"updated": updated}


@app.post("/indices/sync_today")
def sync_today(db: Session = Depends(get_db)):
    """
    手动触发一次增量同步（通常由定时任务自动完成）。
    默认同步区间为 [昨天, 今天]。
    """
    indices = db.query(Index).all()

    log = SyncJobLog(job_type="sync_today", status="running")
    db.add(log)
    db.commit()
    db.refresh(log)

    try:
        # 使用东八区时间
        today = datetime.now(TZ_SHANGHAI).date()
        start_date = today - timedelta(days=1)
        updated, sources = fetch_history_for_indices(
            db, indices, start=start_date, end=today
        )
        log.status = "success"
        log.total_indices = len(indices)
        log.total_rows = int(sum(updated.values()))
        source_counts: dict[str, int] = {}
        for src in sources.values():
            if not src:
                continue
            source_counts[src] = source_counts.get(src, 0) + 1
        if source_counts:
            log.source_summary = ", ".join(f"{k}: {v}" for k, v in source_counts.items())
        
        # 详情格式：indices code：开始日期-截止日期
        detail_parts = []
        for idx in indices:
            if idx.id in updated and updated[idx.id] > 0:
                # 获取该指数的日期范围
                from .fetcher import _get_last_trade_date
                last_date = _get_last_trade_date(db, idx.code)
                first_date = None
                if last_date:
                    # 查询最早日期
                    from sqlalchemy import select, func
                    from .models import IndexPrice
                    stmt = (
                        select(func.min(IndexPrice.trade_date))
                        .where(IndexPrice.index_code == idx.code)
                    )
                    first_date = db.execute(stmt).scalar_one_or_none()
                
                if first_date and last_date:
                    detail_parts.append(f"{idx.code}：{first_date.isoformat()}-{last_date.isoformat()}")
                elif last_date:
                    detail_parts.append(f"{idx.code}：-{last_date.isoformat()}")
        
        log.detail = "；".join(detail_parts) if detail_parts else json.dumps(updated, ensure_ascii=False)
        log.finished_at = now_shanghai()
        db.commit()
    except Exception as exc:
        db.rollback()
        log.status = "error"
        log.detail = f"{type(exc).__name__}: {exc}"
        log.finished_at = now_shanghai()
        db.commit()
        raise

    return {"updated": updated}


@app.post("/indices/recalculate_ma")
def recalculate_ma(
    index_code: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    重新计算并写入均线（30、60、90、120、150、180、360），不写点位差。
    """
    db.info["operator"] = "api"
    try:
        updated_count = recalculate_moving_averages_and_diffs(
            db, index_code=index_code, ma_only=True
        )
        return {
            "success": True,
            "updated_count": updated_count,
            "index_code": index_code if index_code else "all",
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"重新计算均值失败: {type(exc).__name__}: {exc}",
        )


@app.post("/indices/recalculate_ma_diff")
def recalculate_ma_diff(
    index_code: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    根据已有均线与收盘价重新计算并写入点位差（收盘价 - 均线，30/60/90/120/150/180/360）。
    """
    db.info["operator"] = "api"
    try:
        updated_count = recalculate_moving_averages_and_diffs(
            db, index_code=index_code, diff_only=True
        )
        return {
            "success": True,
            "updated_count": updated_count,
            "index_code": index_code if index_code else "all",
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"重新计算均值偏差失败: {type(exc).__name__}: {exc}",
        )


@app.post("/indices/{index_id}/recalculate_ma")
def recalculate_ma_for_index(
    index_id: int,
    db: Session = Depends(get_db),
):
    """
    为单个指数重新计算并写入均线（30、60、90、120、150、180、360），不写点位差。
    """
    db.info["operator"] = "api"
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    
    try:
        updated_count = recalculate_moving_averages_and_diffs(
            db, index_code=index_obj.code, ma_only=True
        )
        return {
            "success": True,
            "updated_count": updated_count,
            "index_code": index_obj.code,
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"重新计算均值失败: {type(exc).__name__}: {exc}",
        )


@app.post("/indices/{index_id}/recalculate_ma_diff")
def recalculate_ma_diff_for_index(
    index_id: int,
    db: Session = Depends(get_db),
):
    """
    为单个指数根据已有均线与收盘价重新计算并写入点位差（收盘价 - 均线，30/60/90/120/150/180/360）。
    """
    db.info["operator"] = "api"
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    
    try:
        updated_count = recalculate_moving_averages_and_diffs(
            db, index_code=index_obj.code, diff_only=True
        )
        return {
            "success": True,
            "updated_count": updated_count,
            "index_code": index_obj.code,
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"重新计算均值偏差失败: {type(exc).__name__}: {exc}",
        )


@app.get("/indices/{index_id}/prices", response_model=PriceQueryResponse)
def get_prices(
    index_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
):
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")

    q = db.query(IndexPrice).filter(IndexPrice.index_code == index_obj.code)
    if start_date:
        q = q.filter(IndexPrice.trade_date >= start_date)
    if end_date:
        q = q.filter(IndexPrice.trade_date <= end_date)
    q = q.order_by(IndexPrice.trade_date.asc())
    rows = q.all()

    prices = [IndexPriceRead.model_validate(row) for row in rows]
    return PriceQueryResponse(index=IndexRead.model_validate(index_obj), prices=prices)


@app.get("/indices/{index_id}/prices/export")
def export_index_prices_excel(
    index_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """
    导出指定指数的行情数据为 Excel 文件。
    支持按 start_date、end_date 筛选日期范围；不传则导出全部。
    """
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")

    q = db.query(IndexPrice).filter(IndexPrice.index_code == index_obj.code)
    if start_date:
        q = q.filter(IndexPrice.trade_date >= start_date)
    if end_date:
        q = q.filter(IndexPrice.trade_date <= end_date)
    q = q.order_by(IndexPrice.trade_date.asc())
    rows = q.all()

    if not rows:
        raise HTTPException(status_code=404, detail="该日期范围内无行情数据")

    # 表头与 IndexPrice 字段对应，使用中文列名；数值列需导出为 Excel 数字格式
    columns = [
        ("trade_date", "日期"),
        ("index_code", "指数代码"),
        ("index_name", "指数名称"),
        ("close", "收盘价"),
        ("change_pct", "涨跌幅(%)"),
        ("pe_ratio", "市盈率"),
        ("ma_30", "MA30"),
        ("ma_60", "MA60"),
        ("ma_90", "MA90"),
        ("ma_120", "MA120"),
        ("ma_150", "MA150"),
        ("ma_180", "MA180"),
        ("ma_360", "MA360"),
        ("ma30_diff", "MA30偏差"),
        ("ma60_diff", "MA60偏差"),
        ("ma90_diff", "MA90偏差"),
        ("ma120_diff", "MA120偏差"),
        ("ma150_diff", "MA150偏差"),
        ("ma180_diff", "MA180偏差"),
        ("ma360_diff", "MA360偏差"),
    ]
    numeric_keys = [k for k, _ in columns if k not in ("trade_date", "index_code", "index_name")]

    def _to_float(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    data = []
    for row in rows:
        data.append({
            "trade_date": row.trade_date.isoformat() if row.trade_date else "",
            "index_code": str(row.index_code) if row.index_code else "",
            "index_name": str(row.index_name) if row.index_name else "",
            "close": _to_float(row.close),
            "change_pct": _to_float(row.change_pct),
            "pe_ratio": _to_float(row.pe_ratio),
            "ma_30": _to_float(row.ma_30),
            "ma_60": _to_float(row.ma_60),
            "ma_90": _to_float(row.ma_90),
            "ma_120": _to_float(row.ma_120),
            "ma_150": _to_float(row.ma_150),
            "ma_180": _to_float(row.ma_180),
            "ma_360": _to_float(row.ma_360),
            "ma30_diff": _to_float(row.ma30_diff),
            "ma60_diff": _to_float(row.ma60_diff),
            "ma90_diff": _to_float(row.ma90_diff),
            "ma120_diff": _to_float(row.ma120_diff),
            "ma150_diff": _to_float(row.ma150_diff),
            "ma180_diff": _to_float(row.ma180_diff),
            "ma360_diff": _to_float(row.ma360_diff),
        })
    df = pd.DataFrame(data)
    # 确保数值列为 float，Excel 中显示为数字格式
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")
    df.columns = [col[1] for col in columns]

    try:
        buf = BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        excel_bytes = buf.getvalue()
        # 设置 Excel 数字格式：涨跌幅为百分比，其余数值为两位小数
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(excel_bytes), read_only=False, data_only=False)
        ws = wb.active
        # 列 4=收盘价, 5=涨跌幅(%), 6=市盈率, 7-13=MA, 14-20=偏差（Excel 列号 1-based）
        for r in range(2, ws.max_row + 1):
            for c in range(4, 21):  # D 到 T
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                if c == 5:  # 涨跌幅(%)
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "0.00"
        out_buf = BytesIO()
        wb.save(out_buf)
        excel_bytes = out_buf.getvalue()
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成 Excel 失败: {e}")

    safe_name = f"行情_{index_obj.code}_{index_obj.name}_{start_date or '起'}_{end_date or '止'}.xlsx"
    for c in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        safe_name = safe_name.replace(c, "_")
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_name)}"},
    )


@app.post("/backtest", response_model=BacktestResponse)
def run_backtest(payload: BacktestRequest, db: Session = Depends(get_db)):
    index_obj = db.query(Index).filter(Index.id == payload.index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    if payload.initial_cash <= 0:
        raise HTTPException(status_code=400, detail="initial_cash 必须大于 0")

    q = db.query(IndexPrice).filter(IndexPrice.index_code == index_obj.code)
    if payload.start_date:
        q = q.filter(IndexPrice.trade_date >= payload.start_date)
    if payload.end_date:
        q = q.filter(IndexPrice.trade_date <= payload.end_date)
    q = q.order_by(IndexPrice.trade_date.asc())
    rows = q.all()
    if not rows:
        raise HTTPException(status_code=404, detail="该日期范围内无行情数据")

    strategy = payload.strategy
    params: dict = {}
    try:
        if strategy == "dca":
            if not payload.dca:
                raise HTTPException(status_code=400, detail="dca 配置不能为空")
            if payload.dca.amount <= 0:
                raise HTTPException(status_code=400, detail="dca.amount 必须大于 0")
            params = {
                "amount": payload.dca.amount,
                "frequency": payload.dca.frequency,
            }
            curve, trades, total_invested = backtest_dca(
                rows=rows,
                initial_cash=payload.initial_cash,
                periodic_amount=payload.dca.amount,
                frequency=payload.dca.frequency,
            )
        elif strategy == "smart_dca":
            if not payload.smart_dca:
                raise HTTPException(status_code=400, detail="smart_dca 配置不能为空")
            params = {
                "base_amount": payload.smart_dca.base_amount,
                "frequency": payload.smart_dca.frequency,
                "step_pct": payload.smart_dca.step_pct,
                "max_multiplier": payload.smart_dca.max_multiplier,
            }
            curve, trades, total_invested = backtest_smart_dca(
                rows=rows,
                initial_cash=payload.initial_cash,
                base_amount=payload.smart_dca.base_amount,
                frequency=payload.smart_dca.frequency,
                step_pct=payload.smart_dca.step_pct,
                max_multiplier=payload.smart_dca.max_multiplier,
            )
        elif strategy == "mean_reversion":
            if not payload.mean_reversion:
                raise HTTPException(status_code=400, detail="mean_reversion 配置不能为空")
            params = {
                "ma_period": payload.mean_reversion.ma_period,
                "buy_threshold_pct": payload.mean_reversion.buy_threshold_pct,
                "sell_threshold_pct": payload.mean_reversion.sell_threshold_pct,
            }
            curve, trades, total_invested = backtest_mean_reversion(
                rows=rows,
                initial_cash=payload.initial_cash,
                ma_period=payload.mean_reversion.ma_period,
                buy_threshold_pct=payload.mean_reversion.buy_threshold_pct,
                sell_threshold_pct=payload.mean_reversion.sell_threshold_pct,
            )
        elif strategy == "mean_reversion_multi":
            if not payload.mean_reversion_multi:
                raise HTTPException(status_code=400, detail="mean_reversion_multi 配置不能为空")
            params = {
                "ma_period": payload.mean_reversion_multi.ma_period,
                "step_pct": payload.mean_reversion_multi.step_pct,
                "order_amount": payload.mean_reversion_multi.order_amount,
            }
            curve, trades, total_invested = backtest_mean_reversion_multi(
                rows=rows,
                initial_cash=payload.initial_cash,
                ma_period=payload.mean_reversion_multi.ma_period,
                step_pct=payload.mean_reversion_multi.step_pct,
                order_amount=payload.mean_reversion_multi.order_amount,
            )
        elif strategy == "ma_diff":
            if not payload.ma_diff:
                raise HTTPException(status_code=400, detail="ma_diff 配置不能为空")
            params = {
                "ma_period": payload.ma_diff.ma_period,
                "buy_diff": payload.ma_diff.buy_diff,
                "sell_diff": payload.ma_diff.sell_diff,
            }
            curve, trades, total_invested = backtest_ma_diff(
                rows=rows,
                initial_cash=payload.initial_cash,
                ma_period=payload.ma_diff.ma_period,
                buy_threshold=payload.ma_diff.buy_diff,
                sell_threshold=payload.ma_diff.sell_diff,
            )
        elif strategy == "trend_following":
            if not payload.trend_following:
                raise HTTPException(status_code=400, detail="trend_following 配置不能为空")
            params = {
                "ma_period": payload.trend_following.ma_period,
            }
            curve, trades, total_invested = backtest_trend_following(
                rows=rows,
                initial_cash=payload.initial_cash,
                ma_period=payload.trend_following.ma_period,
            )
        else:
            raise HTTPException(status_code=400, detail="strategy 不支持")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    metrics_raw = compute_metrics(curve, total_invested)
    metrics = BacktestMetrics(**metrics_raw)
    return BacktestResponse(
        index=IndexRead.model_validate(index_obj),
        strategy=strategy,
        params=params,
        metrics=metrics,
        equity_curve=[BacktestPointRead.model_validate(p.__dict__) for p in curve],
        trades=[BacktestTradeRead.model_validate(t.__dict__) for t in trades],
    )


@app.get("/indices/{index_id}/date_status", response_model=IndexDateStatus)
def get_index_date_status(
    index_id: int,
    db: Session = Depends(get_db),
):
    """
    获取指定指数的日期数据状态，用于日历图展示。
    """
    from sqlalchemy import select, func
    from .models import IndexPrice
    
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    
    # 获取所有价格数据
    stmt = (
        select(IndexPrice)
        .where(IndexPrice.index_code == index_obj.code)
        .order_by(IndexPrice.trade_date.asc())
    )
    price_rows = db.execute(stmt).scalars().all()
    
    # 提取日期列表
    dates = [row.trade_date.isoformat() for row in price_rows]
    
    # 构建日期到价格的映射
    prices_dict = {}
    for row in price_rows:
        date_str = row.trade_date.isoformat()
        prices_dict[date_str] = IndexPriceRead.model_validate(row)
    
    # 获取最早和最晚日期
    first_date = price_rows[0].trade_date if price_rows else None
    last_date = price_rows[-1].trade_date if price_rows else None
    
    return IndexDateStatus(
        index_id=index_obj.id,
        index_code=index_obj.code,
        index_name=index_obj.name,
        dates=dates,
        start_date=index_obj.start_date,
        first_data_date=first_date,
        last_data_date=last_date,
        prices=prices_dict,
    )


@app.get("/indices/date_status", response_model=List[IndexDateStatus])
def get_all_indices_date_status(
    db: Session = Depends(get_db),
):
    """
    获取所有指数的日期数据状态，用于日历图展示。
    """
    indices = db.query(Index).all()
    results = []
    for index_obj in indices:
        try:
            status = get_index_date_status(index_obj.id, db)
            results.append(status)
        except Exception:
            # 跳过出错的指数
            continue
    return results


@app.get("/indices/{index_id}/calendar/{year}", response_model=CalendarDataResponse)
def get_calendar_data(
    index_id: int,
    year: int,
    db: Session = Depends(get_db),
):
    """
    获取指定指数在指定年份的日历图数据。
    返回该年份所有交易日的数据，包括收盘价和涨跌幅。
    """
    index_obj = db.query(Index).filter(Index.id == index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    
    # 计算年份的开始和结束日期
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    
    # 查询该年份的所有价格数据，按日期排序
    prices = (
        db.query(IndexPrice)
        .filter(
            IndexPrice.index_code == index_obj.code,
            IndexPrice.trade_date >= year_start,
            IndexPrice.trade_date <= year_end,
        )
        .order_by(IndexPrice.trade_date.asc())
        .all()
    )
    
    if not prices:
        # 返回空数据
        return CalendarDataResponse(
            index=IndexRead.model_validate(index_obj),
            year=year,
            data=[],
            min_value=0.0,
            max_value=0.0,
        )
    
    # 构建日历数据点，计算涨跌幅
    calendar_data = []
    prev_close = None
    
    for price in prices:
        change_pct = None
        if prev_close is not None and prev_close > 0:
            change_pct = ((price.close - prev_close) / prev_close) * 100
        
        calendar_data.append(
            CalendarDataPoint(
                date=price.trade_date.isoformat(),
                value=price.close,
                change_pct=change_pct,
            )
        )
        prev_close = price.close
    
    # 计算最小值和最大值
    values = [p.value for p in calendar_data]
    min_value = min(values) if values else 0.0
    max_value = max(values) if values else 0.0
    
    return CalendarDataResponse(
        index=IndexRead.model_validate(index_obj),
        year=year,
        data=calendar_data,
        min_value=min_value,
        max_value=max_value,
    )


@app.get("/admin/tables", response_model=List[TableInfo])
def list_tables():
    """
    列出当前数据库中的所有表及行数。
    """
    inspector = inspect(engine)
    table_names = inspector.get_table_names()

    results: List[TableInfo] = []
    with engine.connect() as conn:
        for name in table_names:
            total = conn.execute(text(f"SELECT COUNT(*) FROM {name}")).scalar_one()
            results.append(TableInfo(name=name, row_count=int(total)))
    return results


@app.get("/admin/tables/{table_name}/rows", response_model=TableDataResponse)
def get_table_rows(
    table_name: str,
    offset: int = 0,
    limit: int = 50,
):
    """
    查看指定表的数据，分页返回。
    """
    inspector = inspect(engine)
    valid_tables = set(inspector.get_table_names())
    if table_name not in valid_tables:
        raise HTTPException(status_code=404, detail="表不存在")

    if offset < 0:
        offset = 0
    if limit <= 0:
        limit = 50
    limit = min(limit, 500)

    with engine.connect() as conn:
        total = conn.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar_one()
        result = conn.execute(
            text(f"SELECT * FROM {table_name} LIMIT :limit OFFSET :offset"),
            {"limit": limit, "offset": offset},
        )
        rows = [dict(row) for row in result.mappings().all()]
        columns = list(result.keys())

    return TableDataResponse(
        table=table_name,
        total=int(total),
        offset=offset,
        limit=limit,
        columns=columns,
        rows=rows,
    )


@app.get("/admin/tables/{table_name}/primary_key")
def get_table_primary_key(table_name: str):
    """
    获取指定表的主键列名。
    """
    inspector = inspect(engine)
    valid_tables = set(inspector.get_table_names())
    if table_name not in valid_tables:
        raise HTTPException(status_code=404, detail="表不存在")

    pk_constraint = inspector.get_pk_constraint(table_name)
    pk_columns = pk_constraint.get("constrained_columns", [])
    
    if not pk_columns:
        raise HTTPException(status_code=400, detail="该表没有主键，无法更新")
    
    return {"primary_key": pk_columns[0] if len(pk_columns) == 1 else pk_columns}


@app.put("/admin/tables/{table_name}/rows")
def update_table_row(
    table_name: str,
    request: TableRowUpdateRequest,
    db: Session = Depends(get_db),
):
    primary_key_value = request.primary_key_value
    updates = request.updates
    """
    更新指定表中的一行数据。
    """
    inspector = inspect(engine)
    valid_tables = set(inspector.get_table_names())
    if table_name not in valid_tables:
        raise HTTPException(status_code=404, detail="表不存在")

    # 获取主键列名
    pk_constraint = inspector.get_pk_constraint(table_name)
    pk_columns = pk_constraint.get("constrained_columns", [])
    
    if not pk_columns:
        raise HTTPException(status_code=400, detail="该表没有主键，无法更新")
    
    if len(pk_columns) > 1:
        raise HTTPException(status_code=400, detail="该表有复合主键，暂不支持更新")
    
    pk_column = pk_columns[0]
    
    # 获取表的列信息
    columns = inspector.get_columns(table_name)
    column_info = {col["name"]: col for col in columns}
    
    # 只允许 indices 表进行更新操作
    if table_name != "indices":
        raise HTTPException(
            status_code=403,
            detail=f"只有 indices 表支持修改操作，表 {table_name} 不支持修改"
        )
    
    # indices 表特殊处理：只允许修改 start_date/source 字段
    allowed_fields = {"start_date", "source"}
    invalid_fields = [f for f in updates.keys() if f not in allowed_fields]
    if invalid_fields:
        raise HTTPException(
            status_code=403,
            detail=f"indices 表只允许修改 start_date 字段，不允许修改: {', '.join(invalid_fields)}"
        )
    # 不允许更新主键、code、name、created_at
    forbidden_fields = {pk_column, "code", "name", "created_at"}
    forbidden_updates = [f for f in updates.keys() if f in forbidden_fields]
    if forbidden_updates:
        raise HTTPException(
            status_code=403,
            detail=f"不允许修改以下字段: {', '.join(forbidden_updates)}"
        )
    
    # 构建UPDATE语句
    set_clauses = []
    params = {"pk_value": primary_key_value}
    
    for col_name, col_value in updates.items():
        col_type = column_info[col_name]["type"]
        # 处理NULL值
        if col_value is None:
            set_clauses.append(f"{col_name} = NULL")
        else:
            param_name = f"val_{col_name}"
            set_clauses.append(f"{col_name} = :{param_name}")
            params[param_name] = col_value
    
    if not set_clauses:
        raise HTTPException(status_code=400, detail="没有要更新的字段")
    
    update_sql = f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE {pk_column} = :pk_value"
    
    try:
        with engine.connect() as conn:
            result = conn.execute(text(update_sql), params)
            conn.commit()
            
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail="未找到要更新的行")
            
            # 记录审计日志
            db.info["operator"] = db.info.get("operator", "admin")
            audit_log = DbAuditLog(
                table_name=table_name,
                operation="UPDATE",
                operator=db.info.get("operator", "admin"),
                row_pk=str(primary_key_value),
                detail=json.dumps(updates, ensure_ascii=False, default=str),
            )
            db.add(audit_log)
            db.commit()
            
            return {"success": True, "message": "更新成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


@app.get("/admin/audit_logs", response_model=AuditLogListResponse)
def list_audit_logs(
    table: Optional[str] = None,
    offset: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
):
    """
    查询数据库操作审计日志，按时间倒序，分页返回。
    """
    q = db.query(DbAuditLog).order_by(DbAuditLog.id.desc())
    if table:
        q = q.filter(DbAuditLog.table_name == table)

    total = q.count()
    if offset < 0:
        offset = 0
    if limit <= 0:
        limit = 50
    limit = min(limit, 500)

    items = q.offset(offset).limit(limit).all()
    return AuditLogListResponse(
        total=total,
        offset=offset,
        limit=limit,
        items=[DbAuditLogRead.model_validate(obj) for obj in items],
    )


@app.get("/admin/sync_logs", response_model=SyncJobLogListResponse)
def list_sync_logs(
    job_type: Optional[str] = None,
    offset: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
):
    """
    查询行情同步任务日志，按时间倒序，分页返回。
    """
    q = db.query(SyncJobLog).order_by(SyncJobLog.id.desc())
    if job_type:
        q = q.filter(SyncJobLog.job_type == job_type)

    total = q.count()
    if offset < 0:
        offset = 0
    if limit <= 0:
        limit = 50
    limit = min(limit, 500)

    items = q.offset(offset).limit(limit).all()
    return SyncJobLogListResponse(
        total=total,
        offset=offset,
        limit=limit,
        items=[SyncJobLogRead.model_validate(obj) for obj in items],
    )


# 通知开关（微信公众号）：管理哪些事件会推送模板消息
EVENT_TYPE_DATA_UPDATE = "data_update"


@app.get("/admin/notification_settings", response_model=List[NotificationSettingRead])
def list_notification_settings(db: Session = Depends(get_db)):
    """
    获取所有通知事件开关。若尚未存在「新数据更新」配置，则自动创建（默认值来自环境变量）。
    """
    for event_type, label in [(EVENT_TYPE_DATA_UPDATE, "新数据更新")]:
        obj = db.query(NotificationSetting).filter(
            NotificationSetting.event_type == event_type
        ).one_or_none()
        if obj is None:
            obj = NotificationSetting(
                event_type=event_type,
                enabled=WECHAT_NOTIFY_DATA_UPDATE_DEFAULT if event_type == EVENT_TYPE_DATA_UPDATE else False,
            )
            db.add(obj)
    db.commit()
    items = db.query(NotificationSetting).order_by(NotificationSetting.event_type).all()
    return [NotificationSettingRead.model_validate(x) for x in items]


@app.put("/admin/notification_settings/{event_type}", response_model=NotificationSettingRead)
def update_notification_setting(
    event_type: str,
    body: NotificationSettingUpdate,
    db: Session = Depends(get_db),
):
    """
    开启/关闭指定事件的通知（如 data_update = 新数据更新）。
    """
    obj = db.query(NotificationSetting).filter(
        NotificationSetting.event_type == event_type
    ).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail=f"未知事件类型: {event_type}")
    obj.enabled = body.enabled
    db.commit()
    db.refresh(obj)
    return NotificationSettingRead.model_validate(obj)


@app.get("/admin/wechat/subscribers", response_model=List[WechatSubscriberRead])
def list_wechat_subscribers():
    """
    获取公众号当前订阅者列表（openid + 昵称）。
    需配置 WECHAT_APPID、WECHAT_APPSECRET。
    """
    try:
        subscribers = get_subscriber_list()
        return [WechatSubscriberRead(openid=s["openid"], nickname=s.get("nickname")) for s in subscribers]
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"获取订阅者列表失败: {e}")


@app.post("/admin/wechat/send_custom")
def wechat_send_custom(body: SendCustomMessageRequest):
    """
    向指定订阅者发送自定义模板消息。
    需在 .env 中配置 WECHAT_TEMPLATE_ID_CUSTOM（建议模板含 first、keyword1=内容、keyword2=时间）。
    """
    if not body.openids:
        raise HTTPException(status_code=400, detail="请至少选择一个订阅者")
    if not (body.content or "").strip():
        raise HTTPException(status_code=400, detail="消息内容不能为空")
    try:
        success, failed = send_custom_message(body.openids, body.content.strip())
        return {"sent": success, "failed": failed, "total": len(body.openids)}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"发送失败: {e}")


# 实盘策略
@app.post("/live_strategies", response_model=LiveStrategyRead)
def create_live_strategy(body: LiveStrategyCreate, db: Session = Depends(get_db)):
    index_obj = db.query(Index).filter(Index.id == body.index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    params_json = json.dumps(body.params or {}, ensure_ascii=False)
    emails = ",".join([str(i) for i in (body.email_subscribers or []) if str(i).strip()])
    obj = LiveStrategy(
        name=body.name.strip(),
        index_id=index_obj.id,
        index_code=index_obj.code,
        index_name=index_obj.name,
        strategy=body.strategy,
        params=params_json,
        email_subscribers=emails or None,
        enabled=body.enabled,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return LiveStrategyRead(
        id=obj.id,
        name=obj.name,
        index_id=obj.index_id,
        index_code=obj.index_code,
        index_name=obj.index_name,
        strategy=obj.strategy,
        params=body.params or {},
        email_subscribers=body.email_subscribers or [],
        enabled=obj.enabled,
        last_run_at=obj.last_run_at,
        last_signal_date=obj.last_signal_date,
        last_signal_action=obj.last_signal_action,
        last_signal_price=obj.last_signal_price,
        last_signal_detail=obj.last_signal_detail,
        created_at=obj.created_at,
        updated_at=obj.updated_at,
    )


@app.post("/live_strategies/from_backtest", response_model=LiveStrategyRead)
def create_live_strategy_from_backtest(
    body: LiveStrategyFromBacktest, db: Session = Depends(get_db)
):
    bt = body.backtest
    index_obj = db.query(Index).filter(Index.id == bt.index_id).one_or_none()
    if not index_obj:
        raise HTTPException(status_code=404, detail="指数不存在")
    params = {}
    if bt.strategy == "dca" and bt.dca:
        params = bt.dca.model_dump()
    elif bt.strategy == "smart_dca" and bt.smart_dca:
        params = bt.smart_dca.model_dump()
    elif bt.strategy == "mean_reversion" and bt.mean_reversion:
        params = bt.mean_reversion.model_dump()
    elif bt.strategy == "mean_reversion_multi" and bt.mean_reversion_multi:
        params = bt.mean_reversion_multi.model_dump()
    elif bt.strategy == "ma_diff" and bt.ma_diff:
        params = bt.ma_diff.model_dump()
    elif bt.strategy == "trend_following" and bt.trend_following:
        params = bt.trend_following.model_dump()
    else:
        raise HTTPException(status_code=400, detail="策略参数缺失")
    name = body.name or f"{index_obj.name}-{bt.strategy}"
    obj = LiveStrategy(
        name=name.strip(),
        index_id=index_obj.id,
        index_code=index_obj.code,
        index_name=index_obj.name,
        strategy=bt.strategy,
        params=json.dumps(params, ensure_ascii=False),
        email_subscribers=None,
        enabled=True,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return LiveStrategyRead(
        id=obj.id,
        name=obj.name,
        index_id=obj.index_id,
        index_code=obj.index_code,
        index_name=obj.index_name,
        strategy=obj.strategy,
        params=params,
        email_subscribers=[],
        enabled=obj.enabled,
        last_run_at=obj.last_run_at,
        last_signal_date=obj.last_signal_date,
        last_signal_action=obj.last_signal_action,
        last_signal_price=obj.last_signal_price,
        last_signal_detail=obj.last_signal_detail,
        created_at=obj.created_at,
        updated_at=obj.updated_at,
    )


@app.get("/live_strategies", response_model=list[LiveStrategyRead])
def list_live_strategies(db: Session = Depends(get_db)):
    items = db.query(LiveStrategy).order_by(LiveStrategy.id.asc()).all()
    result = []
    for it in items:
        try:
            params = json.loads(it.params) if it.params else {}
        except Exception:
            params = {}
        emails = [int(e) for e in (it.email_subscribers or "").replace(";", ",").split(",") if e.strip().isdigit()]
        result.append(
            LiveStrategyRead(
                id=it.id,
                name=it.name,
                index_id=it.index_id,
                index_code=it.index_code,
                index_name=it.index_name,
                strategy=it.strategy,
                params=params,
                email_subscribers=emails,
                enabled=it.enabled,
                last_run_at=it.last_run_at,
                last_signal_date=it.last_signal_date,
                last_signal_action=it.last_signal_action,
                last_signal_price=it.last_signal_price,
                last_signal_detail=it.last_signal_detail,
                created_at=it.created_at,
                updated_at=it.updated_at,
            )
        )
    return result


@app.patch("/live_strategies/{strategy_id}", response_model=LiveStrategyRead)
def update_live_strategy(
    strategy_id: int, body: LiveStrategyUpdate, db: Session = Depends(get_db)
):
    obj = db.query(LiveStrategy).filter(LiveStrategy.id == strategy_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="策略不存在")
    if body.name is not None:
        obj.name = body.name.strip()
    if body.enabled is not None:
        obj.enabled = body.enabled
    if body.params is not None:
        obj.params = json.dumps(body.params, ensure_ascii=False)
    if body.email_subscribers is not None:
        obj.email_subscribers = ",".join([str(i) for i in body.email_subscribers if str(i).strip()]) or None
    db.commit()
    db.refresh(obj)
    try:
        params = json.loads(obj.params) if obj.params else {}
    except Exception:
        params = {}
    emails = [int(e) for e in (obj.email_subscribers or "").replace(";", ",").split(",") if e.strip().isdigit()]
    return LiveStrategyRead(
        id=obj.id,
        name=obj.name,
        index_id=obj.index_id,
        index_code=obj.index_code,
        index_name=obj.index_name,
        strategy=obj.strategy,
        params=params,
        email_subscribers=emails,
        enabled=obj.enabled,
        last_run_at=obj.last_run_at,
        last_signal_date=obj.last_signal_date,
        last_signal_action=obj.last_signal_action,
        last_signal_price=obj.last_signal_price,
        last_signal_detail=obj.last_signal_detail,
        created_at=obj.created_at,
        updated_at=obj.updated_at,
    )


@app.delete("/live_strategies/{strategy_id}")
def delete_live_strategy(strategy_id: int, db: Session = Depends(get_db)):
    obj = db.query(LiveStrategy).filter(LiveStrategy.id == strategy_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="策略不存在")
    db.delete(obj)
    db.commit()
    return {"ok": True}


@app.post("/live_strategies/{strategy_id}/run")
def run_live_strategy_now(strategy_id: int, force_email: bool = False):
    results = run_live_strategies(strategy_id=strategy_id, force_email=force_email)
    return {"ok": True, "results": results}


# 联系人管理
_EMAIL_PROVIDERS = {"163", "qq", "gmail"}


def _normalize_email_provider(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = value.strip().lower()
    if not v:
        return None
    if v not in _EMAIL_PROVIDERS:
        raise HTTPException(status_code=400, detail="邮箱类型仅支持 163/qq/gmail")
    return v


@app.get("/contacts", response_model=list[ContactRead])
def list_contacts(db: Session = Depends(get_db)):
    return db.query(Contact).order_by(Contact.id.asc()).all()


@app.post("/contacts", response_model=ContactRead)
def create_contact(body: ContactCreate, db: Session = Depends(get_db)):
    email = body.email.strip()
    existing = db.query(Contact).filter(Contact.email == email).one_or_none()
    if existing:
        existing.nickname = body.nickname.strip()
        existing.phone = (body.phone or "").strip() or None
        existing.email_provider = _normalize_email_provider(body.email_provider)
        existing.email_port = body.email_port
        existing.email_use_ssl = body.email_use_ssl
        existing.email_user = (body.email_user or "").strip() or None
        existing.email_password = (body.email_password or "").strip() or None
        existing.email_sender = (body.email_sender or "").strip() or None
        db.commit()
        db.refresh(existing)
        return existing

    obj = Contact(
        nickname=body.nickname.strip(),
        email=email,
        phone=(body.phone or "").strip() or None,
        email_provider=_normalize_email_provider(body.email_provider),
        email_port=body.email_port,
        email_use_ssl=body.email_use_ssl,
        email_user=(body.email_user or "").strip() or None,
        email_password=(body.email_password or "").strip() or None,
        email_sender=(body.email_sender or "").strip() or None,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@app.patch("/contacts/{contact_id}", response_model=ContactRead)
def update_contact(contact_id: int, body: ContactUpdate, db: Session = Depends(get_db)):
    obj = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="联系人不存在")
    if body.nickname is not None:
        obj.nickname = body.nickname.strip()
    if body.email is not None:
        obj.email = body.email.strip()
    if body.phone is not None:
        obj.phone = body.phone.strip() or None
    if body.email_provider is not None:
        obj.email_provider = _normalize_email_provider(body.email_provider)
    if body.email_port is not None:
        obj.email_port = body.email_port
    if body.email_use_ssl is not None:
        obj.email_use_ssl = body.email_use_ssl
    if body.email_user is not None:
        obj.email_user = body.email_user.strip() or None
    if body.email_password is not None:
        obj.email_password = body.email_password.strip() or None
    if body.email_sender is not None:
        obj.email_sender = body.email_sender.strip() or None
    db.commit()
    db.refresh(obj)
    return obj


@app.delete("/contacts/{contact_id}")
def delete_contact(contact_id: int, db: Session = Depends(get_db)):
    obj = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="联系人不存在")
    db.delete(obj)
    db.commit()
    return {"ok": True}


@app.post("/contacts/{contact_id}/email_test")
def test_contact_email(contact_id: int, db: Session = Depends(get_db)):
    obj = db.query(Contact).filter(Contact.id == contact_id).one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="联系人不存在")
    smtp_config = None
    if obj.email_provider or obj.email_user or obj.email_password or obj.email_sender or obj.email_port:
        smtp_config = {
            "provider": obj.email_provider,
            "port": obj.email_port,
            "use_ssl": obj.email_use_ssl,
            "username": obj.email_user,
            "password": obj.email_password,
            "sender": obj.email_sender or obj.email_user or obj.email,
        }
    safe_config = None
    if smtp_config:
        safe_config = dict(smtp_config)
        if safe_config.get("password"):
            safe_config["password"] = "***"
    _log.info(
        "[测试邮件] contact_id=%s email=%s provider=%s config=%s",
        obj.id,
        obj.email,
        obj.email_provider,
        safe_config,
    )
    ok, msg = test_smtp_connection(smtp_config=smtp_config)
    _log.info("[测试邮件] 连接结果: ok=%s msg=%s", ok, msg)
    if not ok:
        return {"ok": False, "message": msg}
    subject = "实盘策略测试邮件"
    body = "这是一封测试邮件，用于验证 SMTP 配置是否可用。"
    sent = send_email(subject, body, [obj.email], smtp_config=smtp_config)
    _log.info("[测试邮件] 发送结果: sent=%s", sent)
    return {"ok": sent, "message": "已发送" if sent else "发送失败"}


# 邮箱配置
@app.get("/email_config", response_model=Optional[EmailConfigRead])
def get_email_config(db: Session = Depends(get_db)):
    cfg = db.query(EmailConfig).order_by(EmailConfig.id.desc()).first()
    if not cfg:
        return None
    return EmailConfigRead(
        host=cfg.host,
        port=cfg.port,
        username=cfg.username,
        sender=cfg.sender,
        use_tls=cfg.use_tls,
    )


@app.put("/email_config", response_model=EmailConfigRead)
def upsert_email_config(body: EmailConfigUpdate, db: Session = Depends(get_db)):
    cfg = db.query(EmailConfig).order_by(EmailConfig.id.desc()).first()
    if cfg is None:
        cfg = EmailConfig(
            host=body.host.strip(),
            port=body.port,
            username=(body.username or "").strip() or None,
            password=(body.password or "").strip() or None,
            sender=body.sender.strip(),
            use_tls=body.use_tls,
        )
        db.add(cfg)
    else:
        cfg.host = body.host.strip()
        cfg.port = body.port
        cfg.username = (body.username or "").strip() or None
        cfg.password = (body.password or "").strip() or None
        cfg.sender = body.sender.strip()
        cfg.use_tls = body.use_tls
    db.commit()
    db.refresh(cfg)
    return EmailConfigRead(
        host=cfg.host,
        port=cfg.port,
        username=cfg.username,
        sender=cfg.sender,
        use_tls=cfg.use_tls,
    )


# 前端页面入口（单页应用）
@app.get("/", include_in_schema=False)
def frontend_index():
    return FileResponse("frontend/index.html")


# 挂载前端静态文件目录（如有更多静态资源）
app.mount(
    "/static",
    StaticFiles(directory="frontend"),
    name="frontend",
)

